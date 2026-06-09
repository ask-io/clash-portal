import openpyxl


def process_clash_matrix(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Clash Detection Matrix"]

    col_disciplines = {}
    row_disciplines = {}

    # Track horizontal discipline headers (Row 6, starting at Column 5)
    current_disp = None
    for c in range(5, ws.max_column + 1):
        val = ws.cell(row=6, column=c).value
        if val is not None and str(val).strip() != "":
            current_disp = str(val).strip()
        col_disciplines[c] = current_disp

    # Track vertical discipline headers (Column 2, starting at Row 9)
    current_row_disp = None
    for r in range(9, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val is not None and str(val).strip() != "":
            current_row_disp = str(val).strip()
        row_disciplines[r] = current_row_disp

    # --- INDENTATION RESET: This list stays outside the header mapping loops! ---
    clash_records = []

    # Double loop to scan the lower triangle of the matrix
    for r in range(9, ws.max_row + 1):
        diagonal_column = r - 4

        for c in range(5, min(diagonal_column, ws.max_column + 1)):
            cell_value = ws.cell(row=r, column=c).value

           # Skip blanks and 'X' cells
            if cell_value is None or str(cell_value).strip().upper() == 'X':
                continue

            # Convert to a clean string format for safety
            val_str = str(cell_value).strip()

            # Extract when we find valid clash indicators (Accepts both numbers and text 'O')
            if val_str in ['1', '2', '3', 'O', 'o']:
                row_disp_out = col_disciplines.get(c)
                row_el_out = ws.cell(row=8, column=c).value

                col_disp_out = row_disciplines.get(r)
                col_el_out = ws.cell(row=r, column=4).value

                clash_records.append({
                    "Row Discipline": row_disp_out,
                    "Row Element": row_el_out,
                    "Column Discipline": col_disp_out,
                    "Column Element": col_el_out,
                    "Priority": int(val_str) if val_str.isdigit() else val_str.upper()
                })

    # --- INDENTATION RESET: File export happens AFTER the scanning loops finish ---
    if "Clash_List" in wb.sheetnames:
        del wb["Clash_List"]

    ws_list = wb.create_sheet(title="Clash_List")

    # Fixed typo in Column Discipline header string
    headers = ["Row Discipline", "Row Element",
               "Column Discipline", "Column Element", "Priority"]
    ws_list.append(headers)

    # Loop through all found records and append them
    for rec in clash_records:
        ws_list.append([
            rec["Row Discipline"],
            rec["Row Element"],
            rec["Column Discipline"],
            rec["Column Element"],
            rec["Priority"]
        ])

    # Adjust column widths for maximum readability
    ws_list.column_dimensions['A'].width = 22
    ws_list.column_dimensions['B'].width = 38
    ws_list.column_dimensions['C'].width = 22
    ws_list.column_dimensions['D'].width = 38
    ws_list.column_dimensions['E'].width = 12

    wb.save(file_path)
    return clash_records


# --- PRODUCTION RUNNER BLOCK ---
if __name__ == "__main__":
    try:
        total_records = process_clash_matrix("matrix.xlsx")
        print(
            f"Success! Processed the matrix and wrote {len(total_records)} rows into 'Clash_List'.")
    except Exception as e:
        print(f"An execution error occurred: {e}")
