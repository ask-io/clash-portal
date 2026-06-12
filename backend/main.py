from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse as StaticFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import Font, PatternFill, Alignment
import shutil
import os
import io
import openpyxl
from clash_parser import process_clash_matrix, TIER_STYLES, HEADERS, COL_WIDTHS

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND = os.path.join(os.path.dirname(__file__), "frontend")


@app.get("/")
def serve_index():
    return StaticFile(os.path.join(FRONTEND, "index.html"))


@app.get("/app.js")
def serve_js():
    return StaticFile(os.path.join(FRONTEND, "app.js"), media_type="application/javascript")


@app.get("/style.css")
def serve_css():
    return StaticFile(os.path.join(FRONTEND, "style.css"), media_type="text/css")


OUTPUT_FILENAME = "Priority_Clash_Report.xlsx"


def write_tier_sheet(wb, sheet_name, records, style):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = style["tab"]

    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor=style["header_fill"])
    font = Font(bold=True, color=style["header_font"])
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        ws.append([
            rec.get("Row Discipline", ""),
            rec.get("Row Element", ""),
            rec.get("Column Discipline", ""),
            rec.get("Column Element", ""),
            rec.get("Priority", ""),
        ])

    for letter, width in zip(["A", "B", "C", "D", "E"], COL_WIDTHS):
        ws.column_dimensions[letter].width = width


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    temp_location = os.path.join(current_dir, "uploaded_matrix.xlsx")
    final_output_path = os.path.join(current_dir, OUTPUT_FILENAME)

    with open(temp_location, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        all_records, buckets = process_clash_matrix(temp_location)
        print(
            f"Parser complete — {len(all_records)} total records across all tiers")

        # Build a fresh output workbook with one sheet per tier
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default empty sheet

        sheet_map = {
            "1":     ("Tier 1",     TIER_STYLES["1"]),
            "2":     ("Tier 2",     TIER_STYLES["2"]),
            "3":     ("Tier 3",     TIER_STYLES["3"]),
            # Use Tier O style for original O values
            "O":     ("Tier O",     TIER_STYLES["O"]),
            "EMPTY": ("Tier Empty", TIER_STYLES["EMPTY"]),
        }

        for tier_key, (sheet_title, style) in sheet_map.items():
            write_tier_sheet(wb, sheet_title, buckets[tier_key], style)
            print(f"  {sheet_title}: {len(buckets[tier_key])} rows")

        with open(final_output_path, "wb") as f:
            wb.save(f)

        with open(final_output_path, "rb") as f:
            file_bytes = f.read()

        print(f"Sending {len(file_bytes)} bytes to browser")

        return StreamingResponse(
            io.BytesIO(file_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{OUTPUT_FILENAME}"',
                "Content-Length": str(len(file_bytes)),
                "Access-Control-Expose-Headers": "Content-Disposition, Content-Length",
            }
        )

    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})
