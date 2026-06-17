import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TIER_STYLES = {
    "1":     {"tab": "FF0000", "header_fill": "FF0000", "header_font": "FFFFFF"},
    "2":     {"tab": "FF6600", "header_fill": "FF6600", "header_font": "FFFFFF"},
    "3":     {"tab": "FFCC00", "header_fill": "FFCC00", "header_font": "000000"},
    "O":     {"tab": "00B0F0", "header_fill": "00B0F0", "header_font": "FFFFFF"},
    "NA":   {"tab": "70AD47", "header_fill": "70AD47", "header_font": "FFFFFF"},
}

HEADERS = ["Row Discipline", "Row Element",
           "Column Discipline", "Column Element", "Priority", "Clash Rules"]

COL_WIDTHS = [22, 38, 22, 38, 12, 50]


# Fixed start positions structural to the matrix layout and never change (For NORR matrix template)
DATA_ROW_START = 9
DATA_COL_START = 5

# for universal use, we can make these dynamic by scanning the matrix for the last non-empty row and column. This is done in detect_bounds() below.


def detect_bounds(ws):
    data_row_end = DATA_ROW_START
    r = DATA_ROW_START
    while True:
        val = ws.cell(row=r, column=4).value
        if val is None or str(val).strip() == "":
            break
        data_row_end = r
        r += 1

    data_col_end = DATA_COL_START
    c = DATA_COL_START
    while True:
        val = ws.cell(row=8, column=c).value
        if val is None or str(val).strip() == "":
            break
        data_col_end = c
        c += 1

    print(
        f"  Detected matrix bounds: rows {DATA_ROW_START}-{data_row_end}, cols {DATA_COL_START}-{data_col_end}")
    print(
        f"  Matrix size: {data_row_end - DATA_ROW_START + 1} x {data_col_end - DATA_COL_START + 1}")

    return data_row_end, data_col_end


def _write_sheet(wb, sheet_name, records, style):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = style["tab"]

    ws.append(HEADERS)
    fill = PatternFill("solid", fgColor=style["header_fill"])
    font = Font(bold=True, color=style["header_font"])
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        # Use .get() with fallback empty strings to make it completely crash-proof
        priority = rec.get("Priority", "")
        col_elem = rec.get("Column Element", "") or ""
        row_elem = rec.get("Row Element", "") or ""
        clashRules = f"TIER {priority} {col_elem} vs {row_elem}"
        ws.append([
            rec["Row Discipline"],
            rec["Row Element"],
            rec["Column Discipline"],
            rec["Column Element"],
            rec["Priority"],
            clashRules
        ])

    for letter, width in zip(["A", "B", "C", "D", "E", "F"], COL_WIDTHS):
        ws.column_dimensions[letter].width = width


def process_clash_matrix(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True) #If the matrix has formulas, we want the evaluated values, not the formulas themselves. data_only=True ensures that.
    ws = wb["Clash Detection Matrix"]

    # Dynamically detect how large this matrix actually is
    DATA_ROW_END, DATA_COL_END = detect_bounds(ws)

    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        for c in range(DATA_COL_START, DATA_COL_END + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip().upper() == 'O':
                ws.cell(row=r, column=c).value = "O"

    col_disciplines = {}
    current_disp = None
    for c in range(DATA_COL_START, DATA_COL_END + 1):
        val = ws.cell(row=6, column=c).value
        if val is not None and str(val).strip():
            current_disp = str(val).strip()
        col_disciplines[c] = current_disp

    row_disciplines = {}
    current_row_disp = None
    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        val = ws.cell(row=r, column=2).value
        if val is not None and str(val).strip():
            current_row_disp = str(val).strip()
        row_disciplines[r] = current_row_disp

    # Diagonal: row r maps to col (DATA_COL_START + r - DATA_ROW_START)
    def diagonal_col(r):
        return DATA_COL_START + (r - DATA_ROW_START)

    # Scan lower-left triangle only (cols strictly left of the diagonal)
    # The upper-right half is a mirror scanning both would count every clash twice (usually empty, but still).
    buckets = {"1": [], "2": [], "3": [], "O": [], "NA": []}

    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        diag_c = diagonal_col(r)
        for c in range(DATA_COL_START, min(diag_c, DATA_COL_END + 1)):
            cell_value = ws.cell(row=r, column=c).value
            val_cleaned = str(cell_value).strip(
            ).upper() if cell_value is not None else ""

            def make_record(priority_label):
                return {
                    "Row Discipline":    col_disciplines.get(c),
                    "Row Element":       ws.cell(row=8, column=c).value,
                    "Column Discipline": row_disciplines.get(r),
                    "Column Element":    ws.cell(row=r, column=4).value,
                    "Priority":          priority_label,
                    "Clash Rules":       ws.cell(row=r, column=6).value,
                }

            if val_cleaned == "O":
                buckets["O"].append(make_record("O"))

            else:
                num_str = val_cleaned.split(
                    ".")[0] if "." in val_cleaned else val_cleaned
                if num_str in ("1", "2", "3") and num_str.isdigit():
                    buckets[num_str].append(make_record(int(num_str)))

                elif val_cleaned == "" or (cell_value is not None and str(cell_value).strip() == ""):
                    buckets["NA"].append(make_record("NA"))

    for tier, recs in buckets.items():
        print(f"  Tier {tier}: {len(recs)} records")
    print(f"  Total: {sum(len(v) for v in buckets.values())} records")

    # Write sheets
    for old in ["Tier 1", "Tier 2", "Tier 3", "Tier O", "Tier NA", "Clash_List",]:
        if old in wb.sheetnames:
            del wb[old]

    sheet_map = {
        "1":     "Tier 1",
        "2":     "Tier 2",
        "3":     "Tier 3",
        "O":     "Tier O",
        "NA": "Tier NA",
    }

    for tier_key, sheet_title in sheet_map.items():
        _write_sheet(wb, sheet_title, buckets[tier_key], TIER_STYLES[tier_key])

    wb.save(file_path)

    all_records = []
    for tier_key in ("1", "2", "3", "O", "NA"):
        all_records.extend(buckets[tier_key])
    return all_records, buckets


if __name__ == "__main__":
    import os
    test_file = "uploaded_matrix.xlsx"
    if os.path.exists(test_file):
        try:
            records, buckets = process_clash_matrix(test_file)
            print(f"Success! {len(records)} total rows written.")
        except Exception as e:
            print(f"Error: {e}")
    else:
        print(f"Test skipped: '{test_file}' not found.")
